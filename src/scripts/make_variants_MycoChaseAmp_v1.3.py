#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import glob
import os
import sys
import re
import argparse
from datetime import date

# Optional import for report generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# OpenPyXL Illegal Character Regex
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_text(text):
    """Remove illegal characters for Excel."""
    if text is None: return ""
    text = str(text)
    return ILLEGAL_CHARACTERS_RE.sub("", text)

# ----------------------------------------------------------------------
# Arguments Parsing
# ----------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Aggregate VCF results and filter variants (Supports WGS & Targeted).")
parser.add_argument("input_dir", help="Input directory containing .ann.tsv.txt files")
parser.add_argument("output_tsv", help="Output path for variants.tsv")
parser.add_argument("--min_vaf", type=float, default=0.0, help="Minimum VAF threshold")
parser.add_argument("--max_who_group", type=int, default=None, help="Maximum WHO Group to include")
parser.add_argument("--columns", type=str, default=None, help="Comma-separated list of columns to output")
parser.add_argument("--all_columns", action="store_true", help="Output ALL columns in specific detailed order")
parser.add_argument("--pdf", action="store_true", help="Generate PDF reports (Default: False)")
parser.add_argument("--merge_reports", action="store_true", help="Merge TB-Profiler summary if available")
parser.add_argument("--tb_summary", type=str, default=None, help="Path to specific TB-Profiler summary text file")

args = parser.parse_args()

input_dir = args.input_dir
output_tsv = args.output_tsv

if output_tsv.endswith(".variants.tsv"):
    merged_tsv = output_tsv.replace(".variants.tsv", ".gDST_summary.tsv")
else:
    merged_tsv = output_tsv + ".gDST_summary.tsv"

# Directories
RESULTS_DIR = os.path.abspath(os.path.join(input_dir, os.pardir))
LINEAGE_DIR = os.path.join(RESULTS_DIR, "7_spoligotyping")
QC_DIR = os.path.join(RESULTS_DIR, "5_QC")
FASTQC_DIR = os.path.join(RESULTS_DIR, "3_fastqc")
BRACKEN_DIR = os.path.join(RESULTS_DIR, "8_Kraken2")
LD_DIR = os.path.join(RESULTS_DIR, "6_Variant_Call")
REPORT_DIR = os.path.join(RESULTS_DIR, "Reports_PDF")

os.makedirs(REPORT_DIR, exist_ok=True)

# ----------------------------------------------------------------------
# Constants & Headers
# ----------------------------------------------------------------------
out_header = [
    "Sample", "Position", "REF", "ALT", "Depth", "VAF", "Variant_type",
    "Gene", "Gene_ID", "NC_Change", "AA_change", "RAV",
    "Drug", "WHO_group", "Overlap_level", "Note"
]

DRUG_NAMES = [
    "AMK", "BDQ", "CAP", "CFZ", "DLM", "EMB", "ETO", "INH", 
    "KAN", "LFX", "LZD", "MFX", "PZA", "RIF", "STM",
]

# GMA Drug Order
SUMMARY_DRUG_ORDER = [
    "RIF", "INH", "LFX", "MFX", "LZD", "BDQ", "CFZ", "DLM",
    "PZA", "EMB", "AMK", "CAP", "STM", "KAN", "ETO"
]

DRUG_FULL_NAMES = {
    "RIF": "Rifampicin (RIF)", "INH": "Isoniazid (INH)", "LFX": "Levofloxacin (LFX)",
    "MFX": "Moxifloxacin (MFX)", "LZD": "Linezolid (LZD)", "BDQ": "Bedaquiline (BDQ)",
    "CFZ": "Clofazimine (CFZ)", "DLM": "Delamanid (DLM)", "PZA": "Pyrazinamide (PZA)",
    "EMB": "Ethambutol (EMB)", "AMK": "Amikacin (AMK)", "CAP": "Capreomycin (CAP)",
    "STM": "Streptomycin (STM)", "KAN": "Kanamycin (KAN)", "ETO": "Ethionamide (ETO)"
}

# TB-Profiler Name Mapping (Long -> Short) for gDST derivation
TBP_NAME_MAP = {
    "rifampicin": "RIF", "isoniazid": "INH", "ethambutol": "EMB", "pyrazinamide": "PZA",
    "moxifloxacin": "MFX", "levofloxacin": "LFX", "bedaquiline": "BDQ", "delamanid": "DLM",
    "pretomanid": "PA", "linezolid": "LZD", "streptomycin": "STM", "amikacin": "AMK",
    "kanamycin": "KAN", "capreomycin": "CAP", "clofazimine": "CFZ", "ethionamide": "ETO",
    "para-aminosalicylic_acid": "PAS", "cycloserine": "CS"
}

# QC Stats Columns (Added SpikeIn and OnTarget)
QC_RAW_COLS = ["Total_reads_raw", "Total_bases_raw", "Q30_reads_rate_raw",
               "Total_reads_filtered", "Total_bases_filtered", "Q30_reads_rate_filtered"]

MAPPING_STATS_COLS = ["Coverage_depth", "1X_coverage_rate", "50X_coverage_rate", "100X_coverage_rate",
                      "Mapped_reads", "Mapped_rate", "On_Target_Rate", "Mapping_quality", 
                      "Duplicated_reads", "Duplication_rate", 
                      "Read_length", "Insert_size", "Base_quality",
                      "SpikeIn_1_Depth", "SpikeIn_2_Depth", "SpikeIn_3_Depth"]

DEFAULT_SUMMARY_COLUMNS = ["Sample", "Species", "Main_lineage", "Sub_lineage", "Spoligotype", "DR_Type",
                           "Coverage_depth", "On_Target_Rate", "WGS_QC", "RAV_count", "Other_variant_count"]
for d in SUMMARY_DRUG_ORDER: DEFAULT_SUMMARY_COLUMNS.append(f"{d}_Mutation")
for d in SUMMARY_DRUG_ORDER: DEFAULT_SUMMARY_COLUMNS.append(f"{d}_gDST")
DEFAULT_SUMMARY_COLUMNS.append("Large_deletion")

ALL_SUMMARY_COLUMNS = ["Sample", "Species", "Species_Reads", "Species_Fraction",
                       "Main_lineage", "Sub_lineage", "Spoligotype", "DR_Type"]
ALL_SUMMARY_COLUMNS.extend(QC_RAW_COLS)
ALL_SUMMARY_COLUMNS.extend(MAPPING_STATS_COLS)
ALL_SUMMARY_COLUMNS.extend(["WGS_QC", "RAV_count", "Other_variant_count", "Large_deletion"])

for d in SUMMARY_DRUG_ORDER: ALL_SUMMARY_COLUMNS.append(f"{d}_Mutation")
for d in SUMMARY_DRUG_ORDER: ALL_SUMMARY_COLUMNS.append(f"{d}_gDST")

FINAL_COLUMN_ORDER = [
    "Sample", "Species", "Species_Reads", "Species_Fraction",
    "Main_lineage", "Sub_lineage", "TBP_main_lineage", "TBP_sub_lineage",
    "Spoligotype", "TBP_spoligotype",
    "DR_Type", "TBP_drtype",
    "Total_reads_raw", "Total_bases_raw", "Q30_reads_rate_raw",
    "Total_reads_filtered", "Total_bases_filtered", "Q30_reads_rate_filtered",
    "Coverage_depth",
    "On_Target_Rate", "SpikeIn_1_Depth", "SpikeIn_2_Depth", "SpikeIn_3_Depth",
    "1X_coverage_rate", "50X_coverage_rate", "100X_coverage_rate",
    "Mapped_reads", "TBP_num_reads_mapped",
    "Mapped_rate", "TBP_pct_reads_mapped",
    "Mapping_quality", "Duplicated_reads", "Duplication_rate",
    "Read_length", "Insert_size", "Base_quality", "WGS_QC",
    "RAV_count", "Other_variant_count", "TBP_num_dr_variants", "TBP_num_other_variants",
    "Large_deletion",
]
for d in SUMMARY_DRUG_ORDER:
    FINAL_COLUMN_ORDER.extend([f"{d}_gDST", f"TBP_{d}_gDST", f"{d}_Mutation", f"TBP_{d}_Mutation"])
FINAL_COLUMN_ORDER.extend(["TBP_PA_gDST", "TBP_PA_Mutation", "TBP_PAS_gDST", "TBP_PAS_Mutation", "TBP_CS_gDST", "TBP_CS_Mutation"])

# ----------------------------------------------------------------------
# Storage Structures
# ----------------------------------------------------------------------
samples_set = set()
total_variants = {}              
rav_counts = {}                  
per_drug_hits = {}               
summary_front = {}               
pdst_calls = {}                  
drtb_types = {}                  
variants_by_sample = {}          
large_deletion_hits = {}

def init_sample(sample):
    if sample not in total_variants:
        total_variants[sample] = 0
        rav_counts[sample] = 0
        per_drug_hits[sample] = {d: [] for d in SUMMARY_DRUG_ORDER}
        variants_by_sample[sample] = []
        large_deletion_hits[sample] = {}

# ----------------------------------------------------------------------
# DB Loading & Helper Functions
# ----------------------------------------------------------------------
CONVERT_DB = {}
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONVERT_DB_PATH = os.path.join(SCRIPT_DIR, "Variant_convert_DB.xlsx")

def load_convert_db():
    global CONVERT_DB
    if not os.path.exists(CONVERT_DB_PATH):
        print(f"[INFO] Variant_convert_DB.xlsx not found at {CONVERT_DB_PATH}. Skipping.")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(CONVERT_DB_PATH, data_only=True)
        sheet = wb.worksheets[0]
        wb.close()
        idx = {}
        for col in sheet[1]:
            val = str(col.value).strip() if col.value else ""
            if val in ["Start", "Ref", "Alt", "Note"]: idx[val] = col.column - 1
        if not {"Start", "Ref", "Alt", "Note"}.issubset(idx.keys()): return
        db = {}
        for r, row in enumerate(sheet.rows):
            if r == 0: continue
            start, ref, alt, note = row[idx["Start"]].value, row[idx["Ref"]].value, row[idx["Alt"]].value, row[idx["Note"]].value
            if None in (start, ref, alt, note): continue
            start_str = str(int(start)) if isinstance(start, (int, float)) else str(start).strip()
            db[(start_str, str(ref).strip(), str(alt).strip())] = str(note).strip()
        CONVERT_DB = db
    except Exception: pass

def parse_note_to_fields(note):
    if not note or note in (".", "-"): return (None, None, None)
    parts = note.split("_")
    if len(parts) < 2: return (None, None, None)
    gene, change = parts[0].strip(), parts[1].strip()
    nc, aa = (None, None)
    if change.startswith("p."): aa = change
    elif change.startswith(("c.", "n.")): nc = change
    return (gene, nc, aa)

def read_lineage(sample):
    path = os.path.join(LINEAGE_DIR, f"{sample}.lineage.txt")
    if not os.path.exists(path): return ("-", "-")
    try:
        with open(path, "r") as f: lines = [l.strip() for l in f if l.strip()]
        if len(lines) < 2: return ("-", "-")
        cols = lines[-1].split("\t")
        if len(cols) < 8: return ("-", "-")
        main_lin, sub_lin = cols[-2], cols[-1]
        if sub_lin in ("-", "", "None"): sub_lin = main_lin
        return (main_lin, sub_lin)
    except: return ("-", "-")

def read_spoligotype(sample):
    path = os.path.join(LINEAGE_DIR, f"{sample}.spologityping.out")
    if not os.path.exists(path): return "-"
    try:
        with open(path, "r") as f: m = re.findall(r"(\d{15})", f.read())
        return m[-1] if m else "-"
    except: return "-"

def read_extended_qc_stats(sample):
    stats = {k: "-" for k in MAPPING_STATS_COLS}
    stat_path = os.path.join(QC_DIR, f"{sample}_dup_removed.whole.stats")
    whole_mapped = 0; raw_total = 0
    if os.path.exists(stat_path):
        try:
            with open(stat_path, "r") as f:
                for line in f:
                    if not line.startswith("SN"): continue
                    cols = line.strip().split('\t')
                    label = cols[1].replace(":", "")
                    if label == "raw total sequences": raw_total = int(float(cols[2]))
                    elif label == "reads mapped": whole_mapped = int(float(cols[2])); stats["Mapped_reads"] = str(whole_mapped)
                    elif label == "reads duplicated": stats["Duplicated_reads"] = str(int(float(cols[2])))
                    elif label == "average length": stats["Read_length"] = f"{float(cols[2]):.2f}"
                    elif label == "insert size average": stats["Insert_size"] = f"{float(cols[2]):.2f}"
                    elif label == "average quality": stats["Base_quality"] = f"{float(cols[2]):.2f}"
            if raw_total > 0: stats["Mapped_rate"] = f"{(whole_mapped / raw_total) * 100:.2f}"
        except: pass
    
    target_stat_path = os.path.join(QC_DIR, f"{sample}_dup_removed.target.stats")
    if os.path.exists(target_stat_path):
        try:
            with open(target_stat_path, "r") as f:
                for line in f:
                    if "reads mapped:" in line:
                        target_mapped = int(float(line.strip().split('\t')[2]))
                        if whole_mapped > 0: stats["On_Target_Rate"] = f"{(target_mapped / whole_mapped) * 100:.2f}"
        except: pass

    for i in range(1, 4):
        si_path = os.path.join(QC_DIR, f"{sample}_SpikeIn{i}.depth")
        if os.path.exists(si_path):
            try:
                with open(si_path, "r") as f: stats[f"SpikeIn_{i}_Depth"] = f"{float(f.read().strip()):.2f}"
            except: pass

    for k, f in [("Mapping_quality", ".MAPQ"), ("Coverage_depth", ".depth.average")]:
        p = os.path.join(QC_DIR, f"{sample}{f}")
        if os.path.exists(p):
            try:
                with open(p) as fh: stats[k] = f"{float(fh.read().strip()):.2f}"
            except: pass
            
    for c in ["001", "050", "100"]:
        p = os.path.join(QC_DIR, f"{sample}.depth.{c}X")
        if os.path.exists(p):
            try:
                with open(p) as fh: stats[f"{int(c)}X_coverage_rate"] = f"{float(fh.read().strip()):.2f}"
            except: pass
    return stats

def parse_fastqc_pair(r1_path, r2_path):
    total_seqs = 0; total_bases = 0; q30_count = 0
    for path in [r1_path, r2_path]:
        if not os.path.exists(path): continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                in_basic = False; in_q_scores = False
                for line in f:
                    line = line.strip()
                    if not line: continue
                    if line.startswith(">>"):
                        module = line[2:].split("\t")[0]
                        if module == "Basic Statistics": in_basic = True
                        elif module == "Per sequence quality scores": in_q_scores = True
                        elif module == "END_MODULE": in_basic = False; in_q_scores = False
                        continue
                    cols = line.split("\t")
                    if len(cols) < 2: continue
                    if in_basic:
                        if cols[0] == "Total Bases":
                            val_str = cols[1]; multiplier = 1
                            if "Gbp" in val_str: multiplier = 1000000000; val_str = val_str.replace("Gbp", "")
                            elif "Mbp" in val_str: multiplier = 1000000; val_str = val_str.replace("Mbp", "")
                            elif "kbp" in val_str: multiplier = 1000; val_str = val_str.replace("kbp", "")
                            elif "bp" in val_str: val_str = val_str.replace("bp", "")
                            total_bases += int(float(val_str.strip()) * multiplier)
                    elif in_q_scores:
                        if cols[0] == "Quality": continue
                        try:
                            qual = int(cols[0]); count = float(cols[1])
                            total_seqs += int(count)
                            if qual >= 30: q30_count += int(count)
                        except ValueError: continue
        except: pass
    if total_seqs == 0: return ("-", "-", "-")
    q30_rate = (q30_count / total_seqs) * 100
    return (str(total_seqs), f"{total_bases}", f"{q30_rate:.2f}")

def read_bracken_top_species(sample):
    path = os.path.join(BRACKEN_DIR, f"{sample}.S.braken")
    if not os.path.exists(path): return ("-", "-", "-")
    try:
        max_frac = -1.0; top_row = None
        with open(path, "r", encoding="utf-8") as f:
            next(f, None)
            for line in f:
                cols = line.strip().split('\t')
                if len(cols) < 7: continue
                try:
                    frac = float(cols[6])
                    if frac > max_frac: max_frac = frac; top_row = cols
                except: continue
        if top_row: 
            frac_val = float(top_row[6]) * 100
            return (top_row[0], top_row[5], f"{frac_val:.2f}")
        else: return ("-", "-", "-")
    except: return ("-", "-", "-")

def judge_wgs_qc(q30_rate, depth, sp_frac):
    try:
        q30_val = float(q30_rate) if q30_rate not in ("-", "") else 0.0
        depth_val = float(depth) if depth not in ("-", "") else 0.0
        frac_val = float(sp_frac) if sp_frac not in ("-", "") else 0.0
        return "PASS" if (q30_val > 80.0 and depth_val > 30.0 and frac_val > 50.0) else "FAIL"
    except: return "FAIL"

def extract_vaf(hit_str):
    try:
        core_part = hit_str.split("|")[0]
        v = float(core_part.split("_")[-1])
        return v
    except: return 0.0

def transform_hit(hit):
    try:
        v = extract_vaf(hit)
        if v < 10: return None
        return ("*" + hit) if v < 75 else hit
    except: return hit

def determine_dst_call_from_hits(hits):
    if not hits: return "S"
    for h in hits:
        if extract_vaf(h) >= 75: return "R"
    return "S"

def determine_drtb_type(d):
    rif, inh = d.get("RIF", "S"), d.get("INH", "S")
    if inh=="R" and rif=="S": dt = "Hr-TB"
    elif inh=="S" and rif=="R": dt = "RR-TB"
    elif inh=="R" and rif=="R": dt = "MDR-TB"
    else: dt = "Susceptible"
    if dt in ("RR-TB", "MDR-TB") and (d.get("LFX")=="R" or d.get("MFX")=="R"): dt = "pre-XDR-TB"
    if dt == "pre-XDR-TB" and (d.get("LZD")=="R" or d.get("BDQ")=="R"): dt = "XDR-TB"
    return dt

def detect_large_deletions(sample, sample_avg_depth):
    """
    Detect Large Deletions at the amplicon level using MycoChase_Amp_LD.bed.
    - Condition 1: The sample's overall average depth must be >= 100x.
    - Condition 2: Evaluate base-level depth. A base is 'low depth' if <= max(10, 10% of avg depth).
    - Condition 3: An amplicon is dropped out if >= 50% of its bases are 'low depth'.
    - Condition 4: Gene-specific streak rules (1 for small genes, 2 for large genes) 
                   with a blacklist for known systemic false-positive regions.
    """
    gene_to_drug = {"katG": "INH", "pncA": "PZA", "gid": "STM", "ethA": "ETO", "tlyA": "CAP"}
    deletions = {}
    
    # 1. Check overall average depth (Condition 1: Skip detection if < 100x)
    try:
        avg_depth_val = float(sample_avg_depth)
    except:
        avg_depth_val = 0.0
        
    if avg_depth_val < 100.0:
        return deletions

    # 2. Load BED file and classify amplicons based on coordinates
    # Map to the broad target regions extracted in the Bash script
    gene_target_ranges = {
        "katG": (2148889, 2161111),
        "pncA": (2283681, 2294253),
        "gid":  (4402528, 4413202),
        "ethA": (4321004, 4332483),
        "tlyA": (1912940, 1923746)
    }
    
    bed_path = os.path.join(SCRIPT_DIR, "MycoChase_Amp_LD.bed")
    bed_amplicons = {g: [] for g in gene_to_drug.keys()}
    
    if os.path.exists(bed_path):
        with open(bed_path, 'r') as f:
            for line in f:
                parts = line.strip().split('\t')
                if len(parts) < 4: continue
                start, end, name = int(parts[1]), int(parts[2]), parts[3]
                
                # Accurately classify which gene region the amplicon belongs to using coordinates
                for gene, (g_start, g_end) in gene_target_ranges.items():
                    if g_start <= start <= g_end:
                        bed_amplicons[gene].append((start, end, name))
                        break
    
    # 3. Evaluate amplicon depth for each gene
    for gene, drug in gene_to_drug.items():
        amplicons = sorted(bed_amplicons.get(gene, []), key=lambda x: x[0])
        if not amplicons: continue
        
        depth_file = os.path.join(LD_DIR, f"{sample}_{gene}.depth")
        if not os.path.exists(depth_file): continue
        
        # Load per-base depth of the target gene into memory as a dictionary for efficiency
        pos_depths = {}
        try:
            with open(depth_file, 'r') as f:
                for line in f:
                    parts = line.strip().split('\t')
                    if len(parts) >= 3:
                        pos_depths[int(parts[1])] = int(parts[2])
        except:
            continue
            
        # [CORE CHANGE 1] Realistic depth cutoff (10% Rule)
        # To ignore mapping noise or residual reads from low efficiency (e.g. 4-3-40), 
        # use the higher value between 10x and 10% of overall average depth.
        low_depth_threshold = max(10.0, avg_depth_val * 0.10) 
        
        dropout_status = []
        for amp_start, amp_end, amp_name in amplicons:
            amp_length = amp_end - amp_start + 1
            low_depth_count = 0
            
            # [CORE CHANGE 2] Count 'Low Depth Bases' inside the amplicon
            for pos in range(amp_start, amp_end + 1):
                if pos_depths.get(pos, 0) <= low_depth_threshold:
                    low_depth_count += 1
            
            # Even if there's a normal region overlapping with an adjacent amplicon,
            # consider it a dropout if >= 50% of the entire amplicon length is low depth.
            is_dropped = (low_depth_count / amp_length) >= 0.5
            dropout_status.append(is_dropped)
            
        # [CORE CHANGE 3] Dual criteria for consecutive dropout based on gene size
        # Small genes (pncA, gid, tlyA) require only 1 dropout to be deemed a Large Deletion.
        # Large genes (katG, ethA) require 2 consecutive dropouts to prevent false positives.
        min_streak = 1 if gene in ["pncA", "gid", "tlyA"] else 2
        
        # [CORE CHANGE 4] Blacklist for known systemic false-positive regions
        # Exact match of the region string will be ignored to prevent false positives 
        # while preserving actual larger deletions that encompass this region.
        BLACKLIST_REGIONS = {
            "katG": ["2154613-2155073"]
        }
        
        detected_regions = []
        streak_count = 0
        streak_start_idx = -1
        
        for i, status in enumerate(dropout_status):
            if status:
                if streak_count == 0:
                    streak_start_idx = i  # Remember the start index of consecutive dropouts
                streak_count += 1
            else:
                if streak_count >= min_streak:
                    start_coord = amplicons[streak_start_idx][0]
                    end_coord = amplicons[i - 1][1]
                    region_str = f"{start_coord}-{end_coord}"
                    if region_str not in BLACKLIST_REGIONS.get(gene, []):
                        detected_regions.append(region_str)
                streak_count = 0 # Reset streak
                streak_start_idx = -1
                
        # Handle the case where the array ends with a dropout state
        if streak_count >= min_streak:
            start_coord = amplicons[streak_start_idx][0]
            end_coord = amplicons[len(dropout_status) - 1][1]
            region_str = f"{start_coord}-{end_coord}"
            if region_str not in BLACKLIST_REGIONS.get(gene, []):
                detected_regions.append(region_str)
                
        # Format detected regions into a result string
        if detected_regions:
            regions_str = ",".join(detected_regions)
            deletions[drug] = f"{gene}({regions_str})"
            
    return deletions

# ----------------------------------------------------------------------
# 1. Main Processing: Generate variants.tsv
# ----------------------------------------------------------------------
load_convert_db()

# [MODIFIED] Safely collect samples from multiple directories
# Non-MTB samples might lack variant calls or specific QC stats if pipeline branches early.
def add_to_samples(sample_name):
    if sample_name and sample_name != "*":
        samples_set.add(sample_name)
        init_sample(sample_name)

if os.path.exists(QC_DIR):
    for f in glob.glob(os.path.join(QC_DIR, "*_dup_removed.whole.stats")):
        add_to_samples(os.path.basename(f).replace("_dup_removed.whole.stats", ""))

if os.path.exists(BRACKEN_DIR):
    for f in glob.glob(os.path.join(BRACKEN_DIR, "*.S.braken")):
        add_to_samples(os.path.basename(f).replace(".S.braken", ""))

if os.path.exists(FASTQC_DIR):
    for d in os.listdir(FASTQC_DIR):
        if d.endswith("_fastqc") and os.path.isdir(os.path.join(FASTQC_DIR, d)):
            s = d.replace("_Filtered_R1_fastqc", "").replace("_Filtered_R2_fastqc", "") \
                 .replace("_R1_fastqc", "").replace("_R2_fastqc", "") \
                 .replace("_1_fastqc", "").replace("_2_fastqc", "")
            add_to_samples(s)

pattern = os.path.join(input_dir, "*.ann.tsv.txt")
files = sorted(glob.glob(pattern))

for path in files:
    add_to_samples(os.path.basename(path).replace(".ann.tsv.txt", ""))

if not samples_set:
    print(f"[ERROR] No samples detected in {input_dir}, {QC_DIR}, {BRACKEN_DIR}, or {FASTQC_DIR}.")
    sys.exit(1)

with open(output_tsv, "w", newline="", encoding="utf-8") as out_f:
    writer = csv.writer(out_f, delimiter="\t")
    writer.writerow(out_header)

    for path in files:
        sample = os.path.basename(path).replace(".ann.tsv.txt", "")
        # Samples already added to samples_set via add_to_samples
        
        with open(path, "r", encoding="utf-8") as in_f:
            reader = csv.reader(in_f, delimiter="\t"); next(reader, None)
            for row in reader:
                if len(row) < 26: continue
                pos, ref, alt, depth = row[0].strip(), row[1].strip(), row[2].strip(), row[3]
                vaf_raw, var_type = row[4], row[5]; gene, gene_id = row[6], row[7]
                nc, aa = row[8], row[9]; overlap = row[25]
                db_note = CONVERT_DB.get((pos, ref, alt), "-")
                
                if aa and aa not in (".", "-", "") and not aa.startswith("p."): aa = f"p.{aa}"
                
                try: vaf_val = float(vaf_raw); vaf_str = f"{vaf_val:.2f}"
                except: vaf_val = None; vaf_str = vaf_raw
                
                drug_vals = row[10:25]
                sel_drugs, sel_group, sel_group_ints = [], [], []
                rav = "-"
                for idx, v in enumerate(drug_vals):
                    try: iv = int(v)
                    except: continue
                    if iv in (1, 2): rav = "Yes"
                    if 1 <= iv <= 5:
                        sel_drugs.append(DRUG_NAMES[idx]); sel_group.append(str(iv)); sel_group_ints.append(iv)
                if args.min_vaf > 0 and (vaf_val is None or vaf_val < args.min_vaf): continue
                if args.max_who_group is not None:
                    if not sel_group_ints: continue
                    if min(sel_group_ints) > args.max_who_group: continue
                writer.writerow([sample, pos, ref, alt, depth, vaf_str, var_type, gene, gene_id, nc, aa, rav, ";".join(sel_drugs) if sel_drugs else "-", ";".join(sel_group) if sel_group else "-", overlap, db_note])
                variants_by_sample[sample].append({
                    "Position": pos, "REF": ref, "ALT": alt, "Depth": depth, "VAF": vaf_str, "VAF_float": vaf_val,
                    "Variant_type": var_type, "Gene": gene, "Gene_ID": gene_id, "NC_Change": nc, "AA_change": aa,
                    "RAV": rav, "Drug": ";".join(sel_drugs), "WHO_group": ";".join(sel_group), "Overlap_level": overlap,
                    "drug_indices": sel_drugs,
                    "Note": db_note
                })
                total_variants[sample] += 1
                if rav == "Yes": rav_counts[sample] += 1
                annot = aa if aa not in (".", "") else nc
                for idx, v in enumerate(drug_vals):
                    try: iv = int(v)
                    except: continue
                    if iv in (1, 2): 
                        drug = DRUG_NAMES[idx]
                        hit_base = f"{gene}_{annot}_{vaf_str}"
                        
                        if db_note and db_note != "-":
                            hit = f"{hit_base}|{db_note}"
                        else:
                            hit = hit_base
                        
                        if hit not in per_drug_hits[sample][drug]: per_drug_hits[sample][drug].append(hit)

print(f"Written variants: {output_tsv}")

# 2. Merged Summary
if args.columns:
    requested = [c.strip() for c in args.columns.split(",")]
    target_columns = [c for c in requested if c in ALL_SUMMARY_COLUMNS]
    if not target_columns: target_columns = DEFAULT_SUMMARY_COLUMNS
elif args.all_columns: target_columns = ALL_SUMMARY_COLUMNS
else: target_columns = DEFAULT_SUMMARY_COLUMNS

with open(merged_tsv, "w", newline="", encoding="utf-8") as out_m:
    writer = csv.writer(out_m, delimiter="\t"); writer.writerow(target_columns)
    for sample in sorted(samples_set):
        main_lin, sub_lin = read_lineage(sample); spoligo = read_spoligotype(sample)
        qc_stats = read_extended_qc_stats(sample)
        
        # [FIX] Robust FastQC Data Parsing (Handles _R1/_R2 OR _1/_2 naming)
        raw_r1_p1 = os.path.join(FASTQC_DIR, f"{sample}_R1_fastqc", "fastqc_data.txt")
        raw_r2_p1 = os.path.join(FASTQC_DIR, f"{sample}_R2_fastqc", "fastqc_data.txt")
        raw_r1_p2 = os.path.join(FASTQC_DIR, f"{sample}_1_fastqc", "fastqc_data.txt")
        raw_r2_p2 = os.path.join(FASTQC_DIR, f"{sample}_2_fastqc", "fastqc_data.txt")

        if os.path.exists(raw_r1_p1):
             raw_r1, raw_r2 = raw_r1_p1, raw_r2_p1
        else:
             # Fallback to _1/_2 pattern if _R1/_R2 not found
             raw_r1, raw_r2 = raw_r1_p2, raw_r2_p2
             
        r_reads, r_bases, r_q30 = parse_fastqc_pair(raw_r1, raw_r2)
        
        # Filtered data naming is consistent (enforced by gmawgs)
        filt_r1 = os.path.join(FASTQC_DIR, f"{sample}_Filtered_R1_fastqc", "fastqc_data.txt")
        filt_r2 = os.path.join(FASTQC_DIR, f"{sample}_Filtered_R2_fastqc", "fastqc_data.txt")
        f_reads, f_bases, f_q30 = parse_fastqc_pair(filt_r1, filt_r2)
        
        top_species, sp_reads, sp_frac = read_bracken_top_species(sample)
        
        # [MODIFIED] Check if species is MTBC for TSV clearing
        mtbc_members_list = ["Mycobacterium tuberculosis", "Mycobacterium africanum", "Mycobacterium bovis", 
                             "Mycobacterium canettii", "Mycobacterium caprae", "Mycobacterium microti", "Mycobacterium orygis"]
        is_mtb_sample = any(mtb in str(top_species) for mtb in mtbc_members_list)

        wgs_qc_status = judge_wgs_qc(f_q30, qc_stats["Coverage_depth"], sp_frac)
        num_rav = rav_counts.get(sample, 0); num_other = total_variants.get(sample, 0) - num_rav
        
        # Call detect_large_deletions with sample_avg_depth
        ld_results = detect_large_deletions(sample, qc_stats.get("Coverage_depth", 0))
        large_deletion_hits[sample] = ld_results
        ld_summary_str = ";".join(ld_results.values()) if ld_results else "-"
        
        hits_dict = per_drug_hits.get(sample, {}); calls = {}
        for drug in SUMMARY_DRUG_ORDER:
            hits = hits_dict.get(drug, [])
            calls[drug] = determine_dst_call_from_hits(hits)
            if drug in ld_results: calls[drug] = "R"
        
        drtb_type = determine_drtb_type(calls)
        pdst_calls[sample] = calls; drtb_types[sample] = drtb_type
        
        summary_front[sample] = {
            "Species": top_species, "Main_lineage": main_lin, "Sub_lineage": sub_lin,
            "Spoligotype": spoligo, "Seq_Depth": qc_stats["Coverage_depth"], 
            "WGS_QC": wgs_qc_status, "DR_Type": drtb_type,
            "On_Target_Rate": qc_stats["On_Target_Rate"],
            "SpikeIn_1": qc_stats["SpikeIn_1_Depth"],
            "SpikeIn_2": qc_stats["SpikeIn_2_Depth"],
            "SpikeIn_3": qc_stats["SpikeIn_3_Depth"]
        }

        row_data = {
            "Sample": sample, "Main_lineage": main_lin, "Sub_lineage": sub_lin, "Spoligotype": spoligo, 
            "Total_reads_raw": r_reads, "Total_bases_raw": r_bases, "Q30_reads_rate_raw": r_q30,
            "Total_reads_filtered": f_reads, "Total_bases_filtered": f_bases, "Q30_reads_rate_filtered": f_q30,
            "Species": top_species, "Species_Reads": sp_reads, "Species_Fraction": sp_frac,
            "Duplicated_reads": qc_stats["Duplicated_reads"], "Duplication_rate": qc_stats["Duplication_rate"],
            "Mapped_reads": qc_stats["Mapped_reads"], "Mapped_rate": qc_stats["Mapped_rate"],
            "Read_length": qc_stats["Read_length"], "Insert_size": qc_stats["Insert_size"],
            "Base_quality": qc_stats["Base_quality"], "Mapping_quality": qc_stats["Mapping_quality"],
            "Coverage_depth": qc_stats["Coverage_depth"], "1X_coverage_rate": qc_stats["1X_coverage_rate"],
            "50X_coverage_rate": qc_stats["50X_coverage_rate"], "100X_coverage_rate": qc_stats["100X_coverage_rate"],
            "WGS_QC": wgs_qc_status, "RAV_count": str(num_rav), "Other_variant_count": str(num_other), 
            "DR_Type": drtb_type, "Large_deletion": ld_summary_str,
            "On_Target_Rate": qc_stats["On_Target_Rate"],
            "SpikeIn_1_Depth": qc_stats["SpikeIn_1_Depth"],
            "SpikeIn_2_Depth": qc_stats["SpikeIn_2_Depth"],
            "SpikeIn_3_Depth": qc_stats["SpikeIn_3_Depth"]
        }
        
        if not is_mtb_sample:
            # Clear extended QC and variant metrics for non-MTBC
            fields_to_clear = [
                "Duplicated_reads", "Duplication_rate", "Mapped_reads", "Mapped_rate",
                "Read_length", "Insert_size", "Base_quality", "Mapping_quality",
                "Coverage_depth", "1X_coverage_rate", "50X_coverage_rate", "100X_coverage_rate",
                "WGS_QC", "RAV_count", "Other_variant_count", "DR_Type", "Large_deletion",
                "On_Target_Rate", "SpikeIn_1_Depth", "SpikeIn_2_Depth", "SpikeIn_3_Depth"
            ]
            for f in fields_to_clear:
                if f in row_data:
                    row_data[f] = "-"

        for drug in SUMMARY_DRUG_ORDER:
            if not is_mtb_sample:
                row_data[f"{drug}_gDST"] = "-"
                row_data[f"{drug}_Mutation"] = "-"
            else:
                row_data[f"{drug}_gDST"] = calls.get(drug, "S")
                hits = hits_dict.get(drug, [])
                
                # [MODIFIED] Correct logic to filter out low VAF variants (which are returned as None)
                processed_hits = [transform_hit(h) for h in hits]
                valid_hits = [h for h in processed_hits if h is not None] # Filter out None
                
                formatted_hits = []
                for h in valid_hits:
                    parts = h.split("|")
                    core = parts[0]
                    if len(parts) > 1 and parts[1]:
                        formatted_hits.append(f"{core} ({parts[1]})")
                    else:
                        formatted_hits.append(core)
                row_data[f"{drug}_Mutation"] = ";".join(formatted_hits) if formatted_hits else "-"
        
        out_row = [row_data.get(col, "") for col in target_columns]
        writer.writerow(out_row)

print(f"Written merged summary: {merged_tsv}")

# 3. PDF Generation
def generate_pdf_for_sample(sample):
    if not REPORTLAB_AVAILABLE: return

    pdf_path = os.path.join(REPORT_DIR, f"{sample}.report.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=10)
    header_style = ParagraphStyle(name='Header', parent=styles['Heading3'], fontSize=11, spaceAfter=2, spaceBefore=5)
    normal_style = styles['Normal']; small_style = ParagraphStyle(name='Small', parent=styles['Normal'], fontSize=7, leading=8)

    elements.append(Paragraph("GenoMycAnalyzer Analysis Report", title_style))
    elements.append(Spacer(1, 2*mm))
    info = summary_front.get(sample, {})
    species_name = info.get("Species", "-")
    
    mtbc_members = ["Mycobacterium tuberculosis", "Mycobacterium africanum", "Mycobacterium bovis", 
                    "Mycobacterium canettii", "Mycobacterium caprae", "Mycobacterium microti", "Mycobacterium orygis"]
    is_mtb = any(mtb in species_name for mtb in mtbc_members)
    
    # 1. Info & Species
    # [MODIFIED] Check if MTB to display extended QC data, else basic info only
    basic_data = [
        ["Sample Name", sample], 
        ["Analysis Date", date.today().strftime("%Y-%m-%d")]
    ]
    
    if is_mtb:
        basic_data.extend([
            ["Coverage depth", info.get("Seq_Depth", "-")],
            ["On-Target Rate", f"{info.get('On_Target_Rate', '-')}%"],
            ["Spike-In Depth", f"S1:{info.get('SpikeIn_1','-')} / S2:{info.get('SpikeIn_2','-')} / S3:{info.get('SpikeIn_3','-')}"],
            ["QC Status", info.get("WGS_QC", "-")]
        ])

    t_basic = Table(basic_data, colWidths=[35*mm, 55*mm], rowHeights=6*mm)
    t_basic.setStyle(TableStyle([('BACKGROUND', (0, 0), (0, -1), colors.lightgrey), ('GRID', (0, 0), (-1, -1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 8), ('PADDING', (0,0), (-1,-1), 3)]))

    spoligo = info.get("Spoligotype", "-"); 
    if spoligo.isdigit(): spoligo = spoligo.zfill(15)
    species_data = [["Species", species_name], ["Main lineage", info.get("Main_lineage", "-")], ["Sub lineage", info.get("Sub_lineage", "-")], ["Spoligotype", spoligo]]
    t_species = Table(species_data, colWidths=[35*mm, 55*mm], rowHeights=6*mm)
    t_species.setStyle(TableStyle([('BACKGROUND', (0, 0), (0, -1), colors.lightgrey), ('GRID', (0, 0), (-1, -1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 8), ('PADDING', (0,0), (-1,-1), 3)]))

    master_table = Table([[[Paragraph("Basic Information", header_style), t_basic], [Paragraph("Species Identification", header_style), t_species]]], colWidths=[95*mm, 95*mm])
    master_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(master_table); elements.append(Spacer(1, 5*mm))

    # 2. gDST Data Prep (calculated for all, but appended to elements ONLY if is_mtb)
    dst_header = ["Drug", "Status", Paragraph("Gene Target<sup>*</sup> (AA/nc change, VAF, WHO group)", small_style), "Large Deletion"]
    dst_data = [dst_header]
    
    calls = pdst_calls.get(sample, {})
    raw_variants = variants_by_sample.get(sample, [])
    ld_map = large_deletion_hits.get(sample, {})
    resistant_drugs = []; heteroresistance_list = []

    for drug_code in SUMMARY_DRUG_ORDER:
        full_name = DRUG_FULL_NAMES.get(drug_code, drug_code)
        status = calls.get(drug_code, "S")
        if status == "R": resistant_drugs.append(drug_code)
        
        mutations_list = []
        for v in raw_variants:
            var_drugs = v["Drug"].split(";")
            var_groups = v["WHO_group"].split(";")
            is_target = False; grp_val = ""
            if drug_code in var_drugs:
                try: d_idx = var_drugs.index(drug_code); grp_val = var_groups[d_idx] if d_idx < len(var_groups) else "-"; is_target = (grp_val in ("1", "2"))
                except: pass
            if is_target:
                annot = v["AA_change"] if v["AA_change"] not in (".", "") else v["NC_Change"]
                if v["AA_change"] not in (".", "") and not annot.startswith("p."): 
                    annot = f"p.{annot}"
                vaf_float = v['VAF_float'] if v['VAF_float'] else 0.0
                vaf_display = f"{vaf_float:.1f}"
                
                db_note = v.get("Note", "-")
                note_suffix = f" [{db_note}]" if db_note and db_note != "-" else ""
                
                if vaf_float >= 75.0: mutations_list.append(f"<i>{v['Gene']}</i> ({annot}{note_suffix}, {vaf_display}%, {grp_val})")
                elif 10.0 <= vaf_float < 75.0: heteroresistance_list.append(f"<b>{full_name}:</b> <i>{v['Gene']}</i> ({annot}{note_suffix}, {vaf_display}%, {grp_val})")
        
        mutations_str = ", ".join(mutations_list) if mutations_list else "-"
        ld_val = ld_map.get(drug_code, "-")
        dst_data.append([full_name, status, Paragraph(mutations_str, small_style), Paragraph(ld_val, small_style)])

    # Only display gDST Table if species is MTB
    if is_mtb:
        elements.append(Paragraph("Genotypic Drug Susceptibility Test (gDST)", header_style))
        t_dst = Table(dst_data, colWidths=[35*mm, 15*mm, 80*mm, 40*mm], rowHeights=5.5*mm)
        t_dst.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey), ('GRID', (0, 0), (-1, -1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,0), 9), ('ALIGN', (0,0), (-1,-1), 'LEFT'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TEXTCOLOR', (1,1), (1,-1), colors.red), ('PADDING', (0,0), (-1,-1), 2)]))
        for i, row in enumerate(dst_data):
            if i > 0:
                if row[1] == "R": t_dst.setStyle(TableStyle([('TEXTCOLOR', (1, i), (1, i), colors.red)]))
                else: t_dst.setStyle(TableStyle([('TEXTCOLOR', (1, i), (1, i), colors.black)]))
        elements.append(t_dst)
        
        footnote = """<font size=6>
        *The target genes for resistance mutations were derived from the 'Catalogue of mutations in Mycobacterium tuberculosis complex and their association with drug resistance'. (2021, World Health Organization; Lancet Microbe. 2022;3(4):e265). 
        Resistance (R) is reported when mutations are detected in either group 1 or group 2. 
        AA change, amino acids change; VAF, variant allele frequency
        </font>"""
        elements.append(Paragraph(footnote, small_style)); elements.append(Spacer(1, 3*mm))

    # 4. Final Result
    elements.append(Paragraph("Final Result", header_style))
    result_lines = [f"This isolate is positive for <b><i>{species_name}</i></b>."]
    
    if is_mtb:
        if resistant_drugs: result_lines.append(f"The isolate is resistant to <b>{', '.join(resistant_drugs)}</b>.")
        else: result_lines.append("This isolate is <b>Pan-susceptible</b>.")
        
    t_final = Table([[Paragraph("<br/>".join(result_lines), normal_style)]], colWidths=[190*mm])
    t_final.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, colors.black), ('PADDING', (0, 0), (-1, -1), 6), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(t_final); elements.append(Spacer(1, 3*mm))

    # 5. Comment & Auth
    if is_mtb:
        elements.append(Paragraph("Additional Comment", header_style))
        if heteroresistance_list:
            heteroresistance_list = sorted(list(set(heteroresistance_list)))
            comment_text = "<b>Heteroresistance variants detected (10% <= VAF < 75%):</b><br/>" + "<br/>".join(heteroresistance_list)
            comment_content = Paragraph(comment_text, small_style)
        else: comment_content = ""
        t_comment = Table([[comment_content]], colWidths=[190*mm], rowHeights=[20*mm])
        t_comment.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, colors.black), ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('PADDING', (0, 0), (-1, -1), 5)]))
        elements.append(t_comment); elements.append(Spacer(1, 3*mm))

    elements.append(Paragraph("Authorized By", header_style))
    auth_data = [["Reporting Laboratory", "", "Doctor's Name", ""], ["Address", "", "Signature", ""]]
    t_auth = Table(auth_data, colWidths=[40*mm, 55*mm, 40*mm, 55*mm], rowHeights=8*mm)
    t_auth.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.black), ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey), ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey), ('FONTSIZE', (0,0), (-1,-1), 9), ('PADDING', (0,0), (-1,-1), 4), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(t_auth)

    doc.build(elements)
    print(f"[INFO] Written PDF report: {pdf_path}")

if REPORTLAB_AVAILABLE and args.pdf:
    print("[INFO] Generating PDF reports...")
    for s in sorted(samples_set):
        try: generate_pdf_for_sample(s)
        except Exception as e: print(f"[WARN] PDF fail {s}: {e}")
elif args.pdf and not REPORTLAB_AVAILABLE:
    print("[WARN] PDF requested but reportlab not installed.")
else:
    print("[INFO] PDF generation skipped.")

# 4. Excel Generation
if args.merge_reports:
    print("[INFO] Starting report integration (GMA + TB-Profiler)...")
    tb_results_dir = os.path.join(RESULTS_DIR, "9_TBProfiler")
    
    # Find TB-Profiler summary text file
    tb_summary_file = None
    
    if args.tb_summary and os.path.exists(args.tb_summary):
        tb_summary_file = args.tb_summary
        print(f"[INFO] Using specified TB-Profiler summary: {tb_summary_file}")
    else:
        tb_files = glob.glob(os.path.join(tb_results_dir, "*.txt"))
        candidates = [
            f for f in tb_files 
            if ".distance_matrix" not in os.path.basename(f)
            and ".itol" not in os.path.basename(f)
            and ".variants" not in os.path.basename(f)
            and ".dr" not in os.path.basename(f)
            and ".lineage" not in os.path.basename(f)
        ]
        if candidates:
             candidates.sort(key=len)
             tb_summary_file = candidates[0]
             print(f"[INFO] Automatically detected TB-Profiler summary: {tb_summary_file}")
    
    if not tb_summary_file:
        print(f"[WARN] No TB-Profiler summary text file found in {tb_results_dir}")
        print("        Ensure '--run-tbprofiler' was used and 'collate' step finished.")
    else:
        tb_records = {} # { sample: { "TBP_INH_Mutation": "val", "TBP_main_lineage": "val" } }
        
        # 1. Read and Standardize TB-Profiler Data
        try:
            with open(tb_summary_file, "r", encoding="utf-8", errors='replace') as f:
                reader = csv.reader(f, delimiter="\t")
                header_row = next(reader, None)
                
                if header_row and len(header_row) > 1:
                    # Create mapping from Index -> Standardized TBP Column Name
                    tbp_idx_map = {}
                    for i, col_name in enumerate(header_row):
                        if i == 0: continue # Skip sample ID
                        
                        # Logic: Check if it's a drug
                        if col_name in TBP_NAME_MAP:
                            short_drug = TBP_NAME_MAP[col_name]
                            std_name = f"TBP_{short_drug}_Mutation"
                        else:
                            # Metadata (e.g. main_lineage, dr_type)
                            # Convert 'dr_type' -> 'drtype' if needed for consistency, or keep as is
                            # User requested specific list like 'TBP_main_lineage', 'TBP_drtype'
                            # We map common header names to the requested format
                            if col_name == "dr_type": std_name = "TBP_drtype"
                            elif col_name == "main_lineage": std_name = "TBP_main_lineage"
                            elif col_name == "sub_lineage": std_name = "TBP_sub_lineage"
                            elif col_name == "spoligotype": std_name = "TBP_spoligotype"
                            elif col_name == "num_reads_mapped": std_name = "TBP_num_reads_mapped"
                            elif col_name == "pct_reads_mapped": std_name = "TBP_pct_reads_mapped"
                            elif col_name == "num_dr_variants": std_name = "TBP_num_dr_variants"
                            elif col_name == "num_other_variants": std_name = "TBP_num_other_variants"
                            elif col_name == "target_median_depth": continue
                            else:
                                std_name = f"TBP_{col_name}" # Default fallback prefix
                        
                        tbp_idx_map[i] = std_name

                    # Parse Rows
                    for row in reader:
                        if not row: continue
                        sample_id = row[0]
                        tb_records[sample_id] = {}
                        for i, val in enumerate(row):
                            if i in tbp_idx_map:
                                tb_records[sample_id][tbp_idx_map[i]] = val
        except Exception as e:
             print(f"[ERROR] Reading TB-Profiler file failed: {e}")

        # 2. Create Integrated Excel File (.xlsx)
        integrated_xlsx = merged_tsv.replace(".gDST_summary.tsv", ".Integrated_Report.xlsx")
        if integrated_xlsx == merged_tsv: integrated_xlsx += ".xlsx"
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Integrated Report"
            
            # Load GMA rows (Dictionary based for column lookup)
            gma_dicts = []
            with open(merged_tsv, "r", encoding="utf-8") as gma_f:
                gma_reader = csv.reader(gma_f, delimiter="\t")
                gma_header = next(gma_reader, None)
                if gma_header:
                    for row in gma_reader:
                        row_dict = dict(zip(gma_header, row))
                        gma_dicts.append(row_dict)
            
            # --- Row 1: Main Header (Styled) ---
            # Use FINAL_COLUMN_ORDER
            ws.append(FINAL_COLUMN_ORDER)
            
            # [MODIFIED] Conditional Background Colors
            # TBP_ prefix -> #CCFFCC (Light Green)
            # Others      -> #CCECFF (Light Blue)
            fill_tbp = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            fill_default = PatternFill(start_color="CCECFF", end_color="CCECFF", fill_type="solid")
            
            for col_num, col_name in enumerate(FINAL_COLUMN_ORDER, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
                if col_name.startswith("TBP_"):
                    cell.fill = fill_tbp
                else:
                    cell.fill = fill_default

            # --- Row 2+: Data ---
            for gma_row in gma_dicts:
                sample_id = gma_row.get("Sample", "-")
                tbp_row = tb_records.get(sample_id, None) # Get TB-Profiler data for this sample
                
                # [MODIFIED] Check MTBC for Excel fallback clearing
                species_name = gma_row.get("Species", "-")
                mtbc_members_list = ["Mycobacterium tuberculosis", "Mycobacterium africanum", "Mycobacterium bovis", 
                                     "Mycobacterium canettii", "Mycobacterium caprae", "Mycobacterium microti", "Mycobacterium orygis"]
                is_mtb_excel = any(mtb in species_name for mtb in mtbc_members_list)

                final_row_values = []
                
                for col in FINAL_COLUMN_ORDER:
                    val = "-"
                    
                    # Logic 1: Value from GMA
                    if col in gma_row:
                        val = gma_row[col]
                    
                    # Logic 2: Value is TB-Profiler gDST (Derived)
                    elif col.startswith("TBP_") and col.endswith("_gDST"):
                        # Extract drug name (e.g. TBP_INH_gDST -> INH)
                        parts = col.split("_")
                        if len(parts) >= 3:
                            drug_short = parts[1]
                            
                            # Check if Sample exists in TBP
                            if tbp_row is None:
                                val = "-" # [REQ] Missing sample -> "-"
                            else:
                                # Look up mutation column
                                mut_col = f"TBP_{drug_short}_Mutation"
                                mutation = tbp_row.get(mut_col, "-")
                                
                                # Default is Susceptible
                                val = "S"
                                
                                # [MODIFIED] Check individual mutations split by comma
                                if mutation and mutation != "-":
                                    # Split by comma to handle multiple variants
                                    # e.g. "*inhA..., katG..."
                                    mut_parts = mutation.split(',')
                                    
                                    for part in mut_parts:
                                        # Clean whitespace
                                        clean_part = part.strip()
                                        
                                        # If ANY single variant does NOT start with '*', it is Resistance.
                                        # (Ignore variants starting with '*' like '*inhA', 
                                        #  but accept 'katG..._*...' where '*' is in the middle)
                                        if clean_part and not clean_part.startswith("*"):
                                            val = "R"
                                            break
                    
                    # Logic 3: Value is TB-Profiler Raw/Metadata
                    elif tbp_row and col in tbp_row:
                         val = tbp_row[col]
                    
                    # Logic 4: Fallback for missing TBP data (Metadata/Mutations)
                    elif tbp_row is None and col.startswith("TBP_"):
                         val = "-"

                    # [MODIFIED] Force clear TBP and DR fields for non-MTBC
                    if not is_mtb_excel:
                        if col.startswith("TBP_") or col.endswith("_gDST") or col.endswith("_Mutation"):
                            val = "-"

                    final_row_values.append(clean_text(val))
                
                ws.append(final_row_values)

            wb.save(integrated_xlsx)
            print(f"[INFO] Integrated Excel report created: {integrated_xlsx}")

        except ImportError:
            print("[ERROR] 'openpyxl' library is required for Excel merging.")
        except Exception as e:
            print(f"[ERROR] Failed to generate Excel: {e}")

print("[INFO] All tasks completed.")